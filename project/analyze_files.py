import openpyxl
import os

# 分析真味道&好汤面.xlsx
print("=" * 60)
print("分析: 真味道&好汤面.xlsx")
print("=" * 60)

wb = openpyxl.load_workbook(r'c:\Users\宇宙无敌智慧大将军\Desktop\test\data_analysis_agent\project\真味道&好汤面.xlsx')
print(f'Sheet names: {wb.sheetnames}')

for sheet in wb.sheetnames:
    ws = wb[sheet]
    print(f'\n=== {sheet} ===')
    print(f'Rows: {ws.max_row}, Cols: {ws.max_column}')
    for i, row in enumerate(ws.iter_rows(max_row=min(15, ws.max_row), values_only=True)):
        print(f'Row {i+1}: {row}')

print("\n" + "=" * 60)
print("分析: 香爆脆本月目标进度.xlsx")
print("=" * 60)

wb2 = openpyxl.load_workbook(r'c:\Users\宇宙无敌智慧大将军\Desktop\test\data_analysis_agent\project\香爆脆本月目标进度.xlsx')
print(f'Sheet names: {wb2.sheetnames}')

for sheet in wb2.sheetnames:
    ws = wb2[sheet]
    print(f'\n=== {sheet} ===')
    print(f'Rows: {ws.max_row}, Cols: {ws.max_column}')
    for i, row in enumerate(ws.iter_rows(max_row=min(15, ws.max_row), values_only=True)):
        print(f'Row {i+1}: {row}')

print("\n" + "=" * 60)
print("分析: 香爆脆追踪表客户别.xlsx")
print("=" * 60)

wb3 = openpyxl.load_workbook(r'c:\Users\宇宙无敌智慧大将军\Desktop\test\data_analysis_agent\project\香爆脆追踪表客户别.xlsx')
print(f'Sheet names: {wb3.sheetnames}')

for sheet in wb3.sheetnames:
    ws = wb3[sheet]
    print(f'\n=== {sheet} ===')
    print(f'Rows: {ws.max_row}, Cols: {ws.max_column}')
    for i, row in enumerate(ws.iter_rows(max_row=min(15, ws.max_row), values_only=True)):
        print(f'Row {i+1}: {row}')

print("\n" + "=" * 60)
print("分析: 香爆脆追踪表部所别05.31.xlsx")
print("=" * 60)

wb4 = openpyxl.load_workbook(r'c:\Users\宇宙无敌智慧大将军\Desktop\test\data_analysis_agent\project\香爆脆追踪表部所别05.31.xlsx')
print(f'Sheet names: {wb4.sheetnames}')

for sheet in wb4.sheetnames:
    ws = wb4[sheet]
    print(f'\n=== {sheet} ===')
    print(f'Rows: {ws.max_row}, Cols: {ws.max_column}')
    for i, row in enumerate(ws.iter_rows(max_row=min(15, ws.max_row), values_only=True)):
        print(f'Row {i+1}: {row}')
