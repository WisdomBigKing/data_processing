"""
Excel多Sheet数据解析器
支持解析复杂的Excel文件，包含多个工作表、合并单元格等
"""

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass, field
import json
import re


@dataclass
class SheetInfo:
    """工作表信息"""
    name: str
    row_count: int
    col_count: int
    headers: List[str]
    data_preview: List[List[Any]]
    data_types: Dict[str, str]
    summary: str = ""


@dataclass
class ExcelData:
    """Excel文件数据结构"""
    filename: str
    sheet_count: int
    sheets: Dict[str, SheetInfo]
    raw_dataframes: Dict[str, pd.DataFrame] = field(default_factory=dict)
    
    def to_dict(self) -> Dict:
        """转换为字典格式"""
        return {
            "filename": self.filename,
            "sheet_count": self.sheet_count,
            "sheets": {
                name: {
                    "name": info.name,
                    "row_count": info.row_count,
                    "col_count": info.col_count,
                    "headers": info.headers,
                    "data_preview": info.data_preview[:5],  # 只返回前5行预览
                    "data_types": info.data_types,
                    "summary": info.summary
                }
                for name, info in self.sheets.items()
            }
        }
    
    def get_full_data(self, sheet_name: str) -> Optional[pd.DataFrame]:
        """获取指定工作表的完整数据"""
        return self.raw_dataframes.get(sheet_name)


class ExcelParser:
    """Excel文件解析器"""
    
    def __init__(self):
        self.supported_extensions = ['.xlsx', '.xls', '.xlsm']
    
    def parse(self, filepath: str) -> ExcelData:
        """
        解析Excel文件
        
        Args:
            filepath: Excel文件路径
            
        Returns:
            ExcelData: 解析后的数据结构
        """
        # 使用openpyxl读取工作簿信息
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet_names = wb.sheetnames
        
        sheets = {}
        raw_dataframes = {}
        
        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            
            # 跳过空工作表
            if ws.max_row is None or ws.max_row < 1:
                continue
            
            # 使用pandas读取数据
            try:
                df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
                
                # 清理空行空列
                df = self._clean_dataframe(df)
                
                if df.empty:
                    continue
                
                # 智能检测表头行
                header_row = self._detect_header_row(df)
                
                # 提取表头
                if header_row >= 0:
                    headers = [str(h) if pd.notna(h) else f"列{i+1}" 
                              for i, h in enumerate(df.iloc[header_row])]
                    # 重新设置DataFrame的列名
                    df.columns = headers
                    df = df.iloc[header_row + 1:].reset_index(drop=True)
                else:
                    headers = [f"列{i+1}" for i in range(len(df.columns))]
                    df.columns = headers
                
                # 检测数据类型
                data_types = self._detect_data_types(df)
                
                # 生成数据预览
                preview = df.head(10).values.tolist()
                
                # 生成摘要
                summary = self._generate_sheet_summary(df, sheet_name)
                
                sheet_info = SheetInfo(
                    name=sheet_name,
                    row_count=len(df),
                    col_count=len(df.columns),
                    headers=headers,
                    data_preview=preview,
                    data_types=data_types,
                    summary=summary
                )
                
                sheets[sheet_name] = sheet_info
                raw_dataframes[sheet_name] = df
                
            except Exception as e:
                print(f"解析工作表 {sheet_name} 时出错: {e}")
                continue
        
        wb.close()
        
        # 提取文件名
        import os
        filename = os.path.basename(filepath)
        
        return ExcelData(
            filename=filename,
            sheet_count=len(sheets),
            sheets=sheets,
            raw_dataframes=raw_dataframes
        )
    
    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """清理DataFrame，移除全空的行和列"""
        # 移除全空行
        df = df.dropna(how='all')
        # 移除全空列
        df = df.dropna(axis=1, how='all')
        return df.reset_index(drop=True)
    
    def _detect_header_row(self, df: pd.DataFrame) -> int:
        """
        智能检测表头行
        
        Returns:
            表头行索引，如果没有检测到返回-1
        """
        for i in range(min(5, len(df))):  # 只检查前5行
            row = df.iloc[i]
            # 计算非空字符串单元格的比例
            non_empty = sum(1 for v in row if pd.notna(v) and str(v).strip())
            str_count = sum(1 for v in row if pd.notna(v) and isinstance(v, str))
            
            # 如果这一行大部分是非空字符串，可能是表头
            if non_empty > 0 and str_count / max(non_empty, 1) > 0.5:
                # 检查下一行是否有数值数据
                if i + 1 < len(df):
                    next_row = df.iloc[i + 1]
                    num_count = sum(1 for v in next_row 
                                   if pd.notna(v) and (isinstance(v, (int, float)) or 
                                   (isinstance(v, str) and self._is_numeric(v))))
                    if num_count > 0:
                        return i
        return 0  # 默认第一行为表头
    
    def _is_numeric(self, value: str) -> bool:
        """检查字符串是否为数值"""
        try:
            float(str(value).replace(',', '').replace('%', ''))
            return True
        except:
            return False
    
    def _detect_data_types(self, df: pd.DataFrame) -> Dict[str, str]:
        """检测每列的数据类型"""
        types = {}
        for col in df.columns:
            col_data = df[col].dropna()
            if len(col_data) == 0:
                types[str(col)] = "empty"
                continue
            
            # 检查是否为数值
            numeric_count = sum(1 for v in col_data if isinstance(v, (int, float)) or 
                               (isinstance(v, str) and self._is_numeric(v)))
            
            # 检查是否为百分比
            pct_count = sum(1 for v in col_data if isinstance(v, str) and '%' in str(v))
            
            # 检查是否为日期
            date_count = sum(1 for v in col_data if self._is_date(v))
            
            total = len(col_data)
            
            if pct_count / total > 0.5:
                types[str(col)] = "percentage"
            elif date_count / total > 0.5:
                types[str(col)] = "date"
            elif numeric_count / total > 0.5:
                types[str(col)] = "numeric"
            else:
                types[str(col)] = "text"
        
        return types
    
    def _is_date(self, value) -> bool:
        """检查是否为日期"""
        if pd.isna(value):
            return False
        try:
            pd.to_datetime(value)
            return True
        except:
            return False
    
    def _generate_sheet_summary(self, df: pd.DataFrame, sheet_name: str) -> str:
        """生成工作表摘要"""
        summary_parts = []
        summary_parts.append(f"工作表'{sheet_name}'包含{len(df)}行{len(df.columns)}列数据。")
        
        # 统计数值列
        numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
        if numeric_cols:
            summary_parts.append(f"数值列: {', '.join(str(c) for c in numeric_cols[:5])}{'...' if len(numeric_cols) > 5 else ''}")
        
        # 检测可能的分组列（通常是前几列的文本列）
        text_cols = [col for col in df.columns[:3] 
                    if df[col].dtype == 'object' and df[col].nunique() < len(df) * 0.5]
        if text_cols:
            summary_parts.append(f"可能的分组列: {', '.join(str(c) for c in text_cols)}")
        
        return " ".join(summary_parts)
    
    def get_analysis_context(self, excel_data: ExcelData) -> str:
        """
        生成用于LLM分析的上下文描述
        
        Args:
            excel_data: 解析后的Excel数据
            
        Returns:
            str: 上下文描述文本
        """
        context_parts = []
        context_parts.append(f"## Excel文件: {excel_data.filename}")
        context_parts.append(f"包含 {excel_data.sheet_count} 个工作表\n")
        
        for sheet_name, sheet_info in excel_data.sheets.items():
            context_parts.append(f"### 工作表: {sheet_name}")
            context_parts.append(f"- 数据规模: {sheet_info.row_count}行 × {sheet_info.col_count}列")
            context_parts.append(f"- 列名: {', '.join(sheet_info.headers[:10])}{'...' if len(sheet_info.headers) > 10 else ''}")
            context_parts.append(f"- {sheet_info.summary}")
            
            # 添加数据预览
            context_parts.append("- 数据预览(前3行):")
            for i, row in enumerate(sheet_info.data_preview[:3]):
                row_str = " | ".join(str(v)[:20] for v in row[:8])
                context_parts.append(f"  Row {i+1}: {row_str}")
            context_parts.append("")
        
        return "\n".join(context_parts)


def parse_multiple_files(filepaths: List[str]) -> List[ExcelData]:
    """
    解析多个Excel文件
    
    Args:
        filepaths: 文件路径列表
        
    Returns:
        List[ExcelData]: 解析结果列表
    """
    parser = ExcelParser()
    results = []
    
    for filepath in filepaths:
        try:
            data = parser.parse(filepath)
            results.append(data)
        except Exception as e:
            print(f"解析文件 {filepath} 失败: {e}")
    
    return results
