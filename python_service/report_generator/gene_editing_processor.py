"""
基因编辑数据处理器 v4.1

核心需求：
1. 保留所有序号组（不去重，保留原始序号）：
   - 有高亮行：输出所有WT行 + 所有高亮行（完整列）
   - 无高亮行：仅输出所有WT行（简化行，只到H列）
2. 精确提取高亮行中红色标记的20bp序列（原样保留大小写和标红）
3. 不去重、不重新编号
"""

import pandas as pd
import numpy as np
import re
import os
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass, field
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


@dataclass
class SequenceEntry:
    """序列条目数据结构"""
    sequence_id: str              # 序号（如036、037等）
    reference_name: str           # 参考序列名称（如GmACC3HiTom）
    row_number: int               # 原始行号
    depth: int                    # 深度
    percentage_raw: str           # 百分比原始字符串（保留原格式，禁止转换）
    data_type: str                # 数据类型（WT/SNP）
    snp_type: str                 # SNP类型（如C->T）
    full_sequence: str            # 完整序列（保留原始大小写）
    left_sequence: str = ""       # 【v10】原始H列序列
    right_sequence: str = ""      # 【v10】原始I列序列
    is_wt: bool = False           # 是否为野生型
    is_highlighted: bool = False  # 是否为高亮行
    highlight_color: str = ""     # 高亮颜色（yellow/green/blue/pink等）
    has_valid_data: bool = True   # 是否有有效数据
    original_row: List = None     # 原始行数据
    original_row_cells: List = None  # 原始单元格对象（用于提取红色碱基）
    original_sequence_value: Any = None  # 原始序列单元格值（CellRichText或str，保留红色格式）


@dataclass 
class SimplifiedRow:
    """化简后的行数据结构 - v11"""
    sequence_id: str              # 序号（如036、113等）
    reference_name: str           # 参考序列（如GmALS3HiTom）
    row_type: str                 # 行类型：WT 或 SNP（高亮行）
    red_20_bases: str             # 20个红色碱基（保留原始大小写，小写=突变）
    original_row_values: List = None  # 【v11】原始行所有单元格值（完整保留）
    is_highlighted: bool = False  # 是否为高亮行
    highlight_color: str = ""     # 高亮颜色
    original_row_num: int = 0     # 原始行号
    is_simplified: bool = False   # 是否为简化行（无高亮组）


class GeneEditingProcessor:
    """
    基因编辑数据处理器 v4.0
    
    核心逻辑（基于用户需求）：
    1. 保留所有序号组：
       - 有高亮行的组：输出WT行 + 所有高亮行（完整列）
       - 无高亮行的组：仅输出WT行（只保留到H列，后续列删除）
    2. 精确提取高亮行中红色标记的20bp序列（原样保留大小写和标红）
    3. 所有保留行按顺序重新编号为1,2,3...
    4. 不基于单元格颜色判断高亮（仅识别实际高亮标记的行）
    """
    
    def __init__(self):
        self.sequence_pattern = re.compile(r'^(\d+)-ref(\w+)$')
        self.base_pattern = re.compile(r'^[ATCGatcg]+$')
        self.mutation_pattern = re.compile(r'([ATCG])->([ATCG])', re.IGNORECASE)
        
        # 用于存储每个参考序列的目标20碱基（从模板或源文件学习）
        self.target_bases_map = {}
        
    def parse_source_file(self, filepath: str) -> Dict[str, Dict]:
        """
        解析源基因编辑数据文件
        
        Args:
            filepath: 文件路径
            
        Returns:
            List: 按文件顺序的序列组列表 [{seq_id, ref_name, all_entries}, ...]
        """
        # 检查文件是否存在
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"文件不存在: {filepath}")
        
        # 检查文件大小
        file_size = os.path.getsize(filepath)
        if file_size == 0:
            raise ValueError("文件为空")
        
        # 【修复】使用rich_text=True读取，以提取红色标记的碱基
        # 如果rich_text模式失败，尝试普通模式
        try:
            wb = load_workbook(filepath, data_only=False, rich_text=True)
        except Exception as e:
            # 如果rich_text模式失败（如文件损坏或不支持），尝试普通模式
            try:
                wb = load_workbook(filepath, data_only=False)
                print(f"警告: 无法使用rich_text模式读取文件，已切换到普通模式: {e}")
            except Exception as e2:
                raise ValueError(f"无法读取Excel文件，请确保文件是有效的.xlsx格式: {e2}")
        ws = wb.active
        
        # 【修复v14】使用有序列表保存序列组，不再按序号合并，保持原始文件顺序
        sequence_groups = []  # [{seq_id, ref_name, all_entries}, ...]
        current_group = None
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=1), 1):
            # 获取行数据，保留百分比格式
            row_data = []
            for cell in row:
                val = cell.value
                # 如果是百分比格式的数值，转换为百分比字符串
                if isinstance(val, float) and cell.number_format and '%' in cell.number_format:
                    # 根据格式确定小数位数
                    decimal_places = cell.number_format.count('0') - 1
                    if decimal_places < 0:
                        decimal_places = 2
                    pct_val = val * 100
                    val = f"{pct_val:.{decimal_places}f}%"
                row_data.append(val)
            
            # 跳过完全空行
            if all(v is None or str(v).strip() == '' for v in row_data):
                continue
            
            first_cell = str(row_data[0]).strip() if row_data[0] else ""
            
            # 检查是否为序号行（如"001-refGmACC3HiTom"）
            match = self.sequence_pattern.match(first_cell)
            if match:
                # 【修复v14】每次遇到序号行都创建新的组，不合并同名序号
                current_group = {
                    'seq_id': first_cell,  # 完整序号格式
                    'ref_name': match.group(2),
                    'all_entries': []
                }
                sequence_groups.append(current_group)
                continue
            
            # 解析数据行
            if current_group:
                # 【修复】检测所有非白色高亮（黄/绿/蓝/粉等）
                is_highlighted, highlight_color = self._is_row_highlighted(row)
                
                entry = self._parse_data_row(
                    row_data, row_idx, current_group['seq_id'], current_group['ref_name'], 
                    is_highlighted, highlight_color, row
                )
                
                if entry:
                    current_group['all_entries'].append(entry)
        
        wb.close()
        return sequence_groups
    
    def _is_row_highlighted(self, row) -> Tuple[bool, str]:
        """
        检测行是否有非白色高亮背景（修复：检测所有颜色）
        
        Returns:
            Tuple[bool, str]: (是否高亮, 高亮颜色名称)
        """
        for cell in row:
            if cell.fill and cell.fill.start_color:
                color = cell.fill.start_color.rgb
                if color and isinstance(color, str) and len(color) >= 6:
                    try:
                        # 提取RGB值
                        r = int(color[-6:-4], 16)
                        g = int(color[-4:-2], 16)
                        b = int(color[-2:], 16)
                        
                        # 排除白色和近白色 (R,G,B都>240)
                        if r > 240 and g > 240 and b > 240:
                            continue
                        
                        # 排除无色/透明 (00000000)
                        if r == 0 and g == 0 and b == 0 and color.upper() == '00000000':
                            continue
                        
                        # 检测颜色类型（调整阈值以匹配实际Excel颜色）
                        # 黄色: FFFFFF00 (R=255, G=255, B=0)
                        if r > 200 and g > 200 and b < 100:
                            return True, "yellow"
                        # 绿色: FF92D050 (R=146, G=208, B=80) - 放宽r阈值
                        elif g > 180 and g > r and g > b and b < 120:
                            return True, "green"
                        # 蓝色
                        elif b > 150 and b > r and b > g:
                            return True, "blue"
                        # 粉色
                        elif r > 200 and b > 150 and g < 180:
                            return True, "pink"
                        # 其他非白色/非透明背景
                        elif not (r < 50 and g < 50 and b < 50):
                            # 排除黑色/透明(00000000)
                            if r > 80 or g > 80 or b > 80:
                                return True, "other"
                    except:
                        pass
        return False, ""
    
    def _extract_red_text_bases(self, row) -> str:
        """
        从行中提取红色字体标记的碱基序列
        【修复】支持富文本格式，正确提取标红的20个碱基，保留小写突变位点
        """
        try:
            from openpyxl.cell.rich_text import CellRichText, TextBlock
        except ImportError:
            # openpyxl版本不支持rich_text，返回空
            return ""
        
        red_bases = []
        
        for cell in row:
            if cell.value is None:
                continue
            
            # 【修复】处理富文本格式（CellRichText）
            if isinstance(cell.value, CellRichText):
                for part in cell.value:
                    # 检查是否有红色字体
                    is_red = False
                    if hasattr(part, 'font') and part.font and part.font.color:
                        color = part.font.color
                        if hasattr(color, 'rgb') and color.rgb:
                            rgb = color.rgb
                            if isinstance(rgb, str):
                                # 检查是否为红色 (FFFF0000 或 FF0000)
                                if 'FF0000' in rgb.upper():
                                    is_red = True
                    
                    if is_red and hasattr(part, 'text'):
                        text = part.text
                        # 只保留碱基字符，保留原始大小写
                        bases = ''.join(c for c in text if c.upper() in 'ATCG')
                        if bases:
                            red_bases.append(bases)
            
            # 处理普通单元格（整个单元格标红）
            elif cell.font and cell.font.color:
                color = cell.font.color
                is_red = False
                if color.rgb:
                    rgb = color.rgb
                    if isinstance(rgb, str) and 'FF0000' in rgb.upper():
                        is_red = True
                
                if is_red:
                    val = str(cell.value).strip()
                    if re.match(r'^[ATCGatcg]+$', val):
                        red_bases.append(val)
        
        # 返回提取的红色碱基，保留原始大小写
        return ''.join(red_bases) if red_bases else ""
    
    def _parse_data_row(self, row_data: List, row_idx: int, seq_id: str, 
                        ref_name: str, is_highlighted: bool, highlight_color: str, row_cells) -> Optional[SequenceEntry]:
        """
        解析单行数据 - v5.0 重构版
        
        按精确列索引读取字段，正确判断WT/SNP。
        
        Excel列结构:
        - A(0): Sort (序号)
        - B(1): Reads number (深度)
        - C(2): Ratio (百分比)
        - D(3): Left variation type (WT/SNP/1D)
        - E(4): Right variation type (WT/SNP/1D)
        - F(5): Left variation detail (如 C->T, T, -)
        - G(6): Right variation detail (如 G->A, G, -)
        - H(7): Left reads seq (序列)
        - I(8): Right reads seq (序列)
        
        WT判断规则: 仅当D列和E列都为'WT'时才是真正的野生型
        """
        try:
            # === B列(1): 深度 ===
            depth = 0
            if len(row_data) > 1 and row_data[1] is not None:
                try:
                    depth = int(row_data[1])
                except (ValueError, TypeError):
                    pass
            
            # === C列(2): 百分比 - 原样保留 ===
            percentage_raw = ""
            if len(row_data) > 2 and row_data[2] is not None:
                val = row_data[2]
                if isinstance(val, str) and '%' in val:
                    percentage_raw = val
                elif isinstance(val, float) and 0 < val <= 1:
                    pct_value = val * 100
                    percentage_raw = f"{pct_value:.4f}".rstrip('0').rstrip('.') + "%"
                elif isinstance(val, (int, float)):
                    percentage_raw = f"{float(val):.4f}".rstrip('0').rstrip('.') + "%"
            
            # === D列(3): 左侧变异类型 ===
            left_type = ""
            if len(row_data) > 3 and row_data[3] is not None:
                left_type = str(row_data[3]).strip().upper()
            
            # === E列(4): 右侧变异类型 ===
            right_type = ""
            if len(row_data) > 4 and row_data[4] is not None:
                right_type = str(row_data[4]).strip().upper()
            
            # === F列(5): 左侧变异详情 ===
            left_var = ""
            if len(row_data) > 5 and row_data[5] is not None:
                left_var = str(row_data[5]).strip()
            
            # === G列(6): 右侧变异详情 ===
            right_var = ""
            if len(row_data) > 6 and row_data[6] is not None:
                right_var = str(row_data[6]).strip()
            
            # === 确定SNP类型 ===
            snp_type = ""
            left_has_mutation = bool(self.mutation_pattern.search(left_var))
            right_has_mutation = bool(self.mutation_pattern.search(right_var))
            if left_has_mutation and right_has_mutation:
                snp_type = left_var  # 两侧相同时取左侧（高亮行两侧一致）
            elif left_has_mutation:
                snp_type = left_var
            elif right_has_mutation:
                snp_type = right_var
            
            # === 确定数据类型 ===
            # WT: 仅当D列和E列都为'WT'
            if left_type == 'WT' and right_type == 'WT':
                data_type = 'WT'
            elif 'SNP' in (left_type, right_type):
                data_type = 'SNP'
            else:
                data_type = left_type if left_type != 'WT' else right_type
            
            # === H列(7): 左侧序列 ===
            left_seq = ""
            if len(row_data) > 7 and row_data[7] is not None:
                s = str(row_data[7]).strip()
                cleaned = ''.join(c for c in s if c.upper() in 'ATCG')
                if len(cleaned) > 20:
                    left_seq = s
            
            # === I列(8): 右侧序列 ===
            right_seq = ""
            if len(row_data) > 8 and row_data[8] is not None:
                s = str(row_data[8]).strip()
                cleaned = ''.join(c for c in s if c.upper() in 'ATCG')
                if len(cleaned) > 20:
                    right_seq = s
            
            # === 选择包含突变标记（小写字母）的序列 ===
            has_lower_left = any(c.islower() and c.upper() in 'ATCG' for c in left_seq)
            has_lower_right = any(c.islower() and c.upper() in 'ATCG' for c in right_seq)
            
            # 【修复v9】高亮行优先使用H列（左序列），因为红色20bp标记在H列
            # 非高亮行按小写字母判断
            if is_highlighted and left_seq:
                # 高亮行：H列包含红色标记，优先使用
                full_sequence = left_seq
                orig_seq_val = row_data[7] if len(row_data) > 7 else None
            elif has_lower_left:
                full_sequence = left_seq
                orig_seq_val = row_data[7] if len(row_data) > 7 else None
            elif has_lower_right:
                full_sequence = right_seq
                orig_seq_val = row_data[8] if len(row_data) > 8 else None
            else:
                full_sequence = left_seq or right_seq
                orig_seq_val = row_data[7] if (left_seq and len(row_data) > 7) else (row_data[8] if len(row_data) > 8 else None)
            
            # === WT判断: 仅当两侧都是WT才是真正野生型 ===
            is_wt = (left_type == 'WT' and right_type == 'WT')
            has_valid_data = depth > 0
            
            return SequenceEntry(
                sequence_id=seq_id,
                reference_name=ref_name,
                row_number=row_idx,
                depth=depth,
                percentage_raw=percentage_raw,
                data_type=data_type,
                snp_type=snp_type,
                full_sequence=full_sequence,
                left_sequence=left_seq,      # 【v10】保存原始H列
                right_sequence=right_seq,    # 【v10】保存原始I列
                is_wt=is_wt,
                is_highlighted=is_highlighted,
                highlight_color=highlight_color,
                has_valid_data=has_valid_data,
                original_row=row_data,
                original_row_cells=list(row_cells),
                original_sequence_value=orig_seq_val
            )
            
        except Exception as e:
            print(f"解析行 {row_idx} 失败: {e}")
            return None
    
    def simplify_data(self, sequence_groups: List[Dict], 
                      target_bases_map: Dict[str, str] = None) -> List[SimplifiedRow]:
        """
        化简数据 - v4.1
        
        规则：
        1. 保留所有序号组（不去重，保留原始序号）：
           - 有高亮行：输出所有WT行 + 所有高亮行（完整列，含完整序列）
           - 无高亮行：仅输出所有WT行（简化行，只保留到H列）
        2. 20bp红色序列原样保留（含大小写和标红状态）
        3. 百分比原样保留（禁止转换/四舍五入）
        4. 不去重、不重新编号，保留原始数据
        5. 【修复v14】保持原始文件顺序，不对序号排序
        """
        results = []
        
        # 【修复v14】sequence_groups是有序列表，直接遍历保持原始文件顺序
        for group in sequence_groups:
            seq_id = group['seq_id']
            ref_name = group['ref_name']
            all_entries = group['all_entries']
            
            # 找出所有高亮行
            highlighted_entries = [e for e in all_entries if e.is_highlighted]
            has_highlights = len(highlighted_entries) > 0
            
            # 找所有WT行（不去重，全部保留）
            wt_entries = [e for e in all_entries if e.is_wt]
            
            if has_highlights:
                # === 有高亮行的组：输出所有WT行 + 所有高亮行（完整列） ===
                
                # 1. 输出所有WT行
                for wt_entry in wt_entries:
                    results.append(SimplifiedRow(
                        sequence_id=seq_id,
                        reference_name=ref_name,
                        row_type="WT",
                        red_20_bases="-",
                        original_row_values=wt_entry.original_row,  # 【v11】保存完整原始行
                        is_highlighted=False,
                        highlight_color="",
                        original_row_num=wt_entry.row_number,
                        is_simplified=False
                    ))
                
                # 2. 输出所有高亮行
                for hl_entry in highlighted_entries:
                    red_bases = ""
                    if hl_entry.original_row_cells:
                        red_bases, _, _ = self._extract_red_bases_from_sequence(
                            hl_entry.full_sequence, hl_entry.original_row_cells
                        )
                    
                    results.append(SimplifiedRow(
                        sequence_id=seq_id,
                        reference_name=ref_name,
                        row_type="SNP",
                        red_20_bases=red_bases if red_bases else "-",
                        original_row_values=hl_entry.original_row,  # 【v11】保存完整原始行
                        is_highlighted=True,
                        highlight_color=hl_entry.highlight_color,
                        original_row_num=hl_entry.row_number,
                        is_simplified=False
                    ))
            else:
                # === 无高亮行的组：仅输出所有WT行（简化行） ===
                for wt_entry in wt_entries:
                    results.append(SimplifiedRow(
                        sequence_id=seq_id,
                        reference_name=ref_name,
                        row_type="WT",
                        red_20_bases="-",
                        original_row_values=wt_entry.original_row,  # 【v11】保存完整原始行
                        is_highlighted=False,
                        highlight_color="",
                        original_row_num=wt_entry.row_number,
                        is_simplified=True  # 标记为简化行
                    ))
        
        return results
    
    def _extract_red_bases_from_sequence(self, sequence: str, original_row) -> tuple:
        """
        从原始单元格中直接提取红色字体标记的20bp文本
        【修复v12】直接复制红色标记的文本，一模一样保留小写字母
        返回 (red_20bp, start_pos, end_pos)
        """
        # 【优先】方法1：直接从H列富文本中提取红色标记的文本（原样复制）
        try:
            from openpyxl.cell.rich_text import CellRichText
            if original_row and len(original_row) > 7:
                h_cell = original_row[7]
                if h_cell and hasattr(h_cell, 'value') and h_cell.value:
                    red_text, red_start = self._extract_red_from_cell_with_position(h_cell)
                    if red_text:
                        # 直接返回红色标记的文本，不做任何处理
                        return (red_text, red_start, red_start + len(red_text))
            
            # 尝试I列
            if original_row and len(original_row) > 8:
                i_cell = original_row[8]
                if i_cell and hasattr(i_cell, 'value') and i_cell.value:
                    red_text, red_start = self._extract_red_from_cell_with_position(i_cell)
                    if red_text:
                        return (red_text, red_start, red_start + len(red_text))
        except ImportError:
            pass
        except Exception as e:
            pass
        
        # 方法2：备用 - 如果无法从富文本提取，使用小写字母位置
        if sequence:
            lowercase_positions = [i for i, c in enumerate(sequence) if c.islower() and c.upper() in 'ATCG']
            if lowercase_positions:
                mutation_pos = lowercase_positions[0]
                start = max(0, mutation_pos - 10)
                end = start + 20
                if end > len(sequence):
                    end = len(sequence)
                    start = max(0, end - 20)
                segment = sequence[start:end]
                return (segment, start, end)
        
        return ("", -1, -1)
    
    def _is_red_color(self, rgb_str: str) -> bool:
        """
        判断RGB颜色是否为红色
        【修复v13】正确区分红色(FFFF0000)和黑色(FF000000)
        """
        if not rgb_str:
            return False
        rgb = rgb_str.upper()
        
        # 格式: AARRGGBB (8位) 或 RRGGBB (6位)
        try:
            if len(rgb) == 8:
                # AARRGGBB格式
                r = int(rgb[2:4], 16)
                g = int(rgb[4:6], 16)
                b = int(rgb[6:8], 16)
            elif len(rgb) == 6:
                # RRGGBB格式
                r = int(rgb[0:2], 16)
                g = int(rgb[2:4], 16)
                b = int(rgb[4:6], 16)
            else:
                return False
            
            # 红色: R值高(>180), G和B值低(<100)
            return r > 180 and g < 100 and b < 100
        except:
            return False
    
    def _extract_red_from_cell_with_position(self, cell) -> tuple:
        """
        从单个单元格提取红色字体的文本及其起始位置
        返回 (red_text, start_position)
        """
        try:
            from openpyxl.cell.rich_text import CellRichText
        except ImportError:
            return ("", -1)
        
        # 处理富文本
        if isinstance(cell.value, CellRichText):
            current_pos = 0
            red_start = -1
            red_text_parts = []
            
            for part in cell.value:
                part_text = part.text if hasattr(part, 'text') else str(part)
                is_red = False
                
                if hasattr(part, 'font') and part.font:
                    font = part.font
                    if hasattr(font, 'color') and font.color:
                        color = font.color
                        if hasattr(color, 'rgb') and color.rgb:
                            rgb = str(color.rgb)
                            is_red = self._is_red_color(rgb)
                
                if is_red and part_text:
                    if red_start == -1:
                        red_start = current_pos
                    red_text_parts.append(part_text)
                
                current_pos += len(part_text) if part_text else 0
            
            if red_text_parts:
                return (''.join(red_text_parts), red_start)
        
        # 处理普通单元格
        elif cell.font and cell.font.color:
            color = cell.font.color
            if color.rgb:
                rgb = str(color.rgb)
                if self._is_red_color(rgb):
                    val = str(cell.value) if cell.value else ""
                    return (val, 0)
        
        return ("", -1)
    
    def _extract_red_from_cell(self, cell) -> str:
        """
        从单个单元格提取红色字体的文本
        支持富文本和普通单元格
        """
        try:
            from openpyxl.cell.rich_text import CellRichText
        except ImportError:
            return ""
        
        red_parts = []
        
        # 处理富文本
        if isinstance(cell.value, CellRichText):
            for part in cell.value:
                is_red = False
                # 检查TextBlock的字体颜色
                if hasattr(part, 'font') and part.font:
                    font = part.font
                    if hasattr(font, 'color') and font.color:
                        color = font.color
                        if hasattr(color, 'rgb') and color.rgb:
                            rgb = str(color.rgb).upper()
                            # 红色: FF0000 或 FFFF0000
                            if 'FF0000' in rgb:
                                is_red = True
                
                if is_red and hasattr(part, 'text') and part.text:
                    red_parts.append(part.text)
        
        # 处理普通单元格（整个单元格红色）
        elif cell.font and cell.font.color:
            color = cell.font.color
            if color.rgb:
                rgb = str(color.rgb).upper()
                if 'FF0000' in rgb:
                    val = str(cell.value) if cell.value else ""
                    if val:
                        red_parts.append(val)
        
        return ''.join(red_parts)
    
    def generate_simplified_excel(self, results: List[SimplifiedRow], output_path: str) -> str:
        """
        生成化简后的Excel文件 - v11
        
        规则：
        1. 保留原始行的所有列数据
        2. 在第一列添加提取的20bp（红色标记）
        3. 在第二列添加类型标记（WT/SNP）
        4. 在第三列添加高亮颜色
        5. 后续列为原始数据的所有列
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "化简数据"
        
        # 样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="FF4472C6", end_color="FF4472C6", fill_type="solid")
        color_fills = {
            "yellow": PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid"),
            "green": PatternFill(start_color="FF90EE90", end_color="FF90EE90", fill_type="solid"),
            "blue": PatternFill(start_color="FFADD8E6", end_color="FFADD8E6", fill_type="solid"),
            "pink": PatternFill(start_color="FFFFB6C1", end_color="FFFFB6C1", fill_type="solid"),
            "other": PatternFill(start_color="FFE0E0E0", end_color="FFE0E0E0", fill_type="solid"),
        }
        red_font = Font(bold=True, color="FF0000")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 【v11】确定原始数据的列数
        max_orig_cols = 0
        for row in results:
            if row.original_row_values:
                max_orig_cols = max(max_orig_cols, len(row.original_row_values))
        
        # 【v11】表头：序号 + 红色20bp + 类型 + 高亮颜色 + 原始数据列
        headers = ["序号", "红色20碱基", "类型", "高亮颜色"]
        # 原始数据列从第1列开始（跳过第0列的行号）
        for i in range(1, max_orig_cols):
            headers.append(f"原列{i}")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        
        # 写入数据
        for row_idx, row in enumerate(results, 2):
            is_snp_row = row.row_type == "SNP"
            hl_color = row.highlight_color or ""
            fill = color_fills.get(hl_color, None) if is_snp_row else None
            
            # 列1: 序号（使用完整序号格式如"001-refGmACC3HiTom"）
            cell = ws.cell(row=row_idx, column=1, value=row.sequence_id)
            cell.border = thin_border
            if fill:
                cell.fill = fill
            
            # 列2: 红色20碱基
            cell = ws.cell(row=row_idx, column=2, value=row.red_20_bases)
            cell.border = thin_border
            if is_snp_row and row.red_20_bases != "-":
                cell.font = red_font
            if fill:
                cell.fill = fill
            
            # 列3: 类型
            cell = ws.cell(row=row_idx, column=3, value=row.row_type)
            cell.border = thin_border
            if fill:
                cell.fill = fill
            
            # 列4: 高亮颜色
            cell = ws.cell(row=row_idx, column=4, value=hl_color if hl_color else "-")
            cell.border = thin_border
            if fill:
                cell.fill = fill
            
            # 【v11】从第5列开始写入原始数据（跳过原始第0列的行号）
            if row.original_row_values:
                for orig_col_idx in range(1, len(row.original_row_values)):
                    orig_val = row.original_row_values[orig_col_idx]
                    cell = ws.cell(row=row_idx, column=4 + orig_col_idx, value=orig_val)
                    cell.border = thin_border
                    if fill:
                        cell.fill = fill
        
        # 【v11】动态调整列宽
        ws.column_dimensions['A'].width = 22  # 序号（完整格式如001-refGmACC3HiTom）
        ws.column_dimensions['B'].width = 25  # 红色20碱基
        ws.column_dimensions['C'].width = 8   # 类型
        ws.column_dimensions['D'].width = 10  # 高亮颜色
        # 其余列使用默认宽度或根据内容自动调整
        for col_idx in range(5, 5 + max_orig_cols):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = 15
        
        wb.save(output_path)
        return output_path
    
    def validate_results(self, results: List[SimplifiedRow]) -> Dict[str, Any]:
        """验证化简结果 - v4.0"""
        issues = []
        warnings = []
        
        snp_rows = [r for r in results if r.row_type == "SNP"]
        wt_rows = [r for r in results if r.row_type == "WT"]
        simplified_rows = [r for r in results if r.is_simplified]
        full_rows = [r for r in results if not r.is_simplified]
        
        # 检查SNP行的红色碱基
        for row in snp_rows:
            if row.red_20_bases == "-" or not row.red_20_bases:
                warnings.append(f"序号 {row.sequence_id}: SNP行缺少红色20碱基")
            elif len(row.red_20_bases) < 10:
                warnings.append(f"序号 {row.sequence_id}: 红色碱基长度不足 ({len(row.red_20_bases)})")
        
        # 统计序号组
        seq_ids = set(r.sequence_id for r in results)
        
        return {
            "valid": len(issues) == 0,
            "issues": issues,
            "warnings": warnings,
            "total_rows": len(results),
            "total_sequence_groups": len(seq_ids),
            "wt_rows": len(wt_rows),
            "snp_rows": len(snp_rows),
            "simplified_groups": len(simplified_rows),
            "full_groups": len(set(r.sequence_id for r in full_rows))
        }
    
    def process_file(self, input_path: str, output_path: str = None,
                     target_bases_map: Dict[str, str] = None) -> Tuple[List[SimplifiedRow], str, Dict]:
        """
        处理基因编辑数据文件的完整流程
        
        Args:
            input_path: 输入文件路径
            output_path: 输出文件路径（可选）
            target_bases_map: 参考序列到20个目标碱基的映射（可选）
            
        Returns:
            Tuple: (化简结果列表, 输出文件路径, 验证结果)
        """
        # 解析文件
        entries_by_id = self.parse_source_file(input_path)
        
        # 化简数据
        simplified_results = self.simplify_data(entries_by_id, target_bases_map)
        
        # 验证结果
        validation = self.validate_results(simplified_results)
        
        # 生成输出文件
        if output_path is None:
            base_name = os.path.splitext(os.path.basename(input_path))[0]
            output_dir = os.path.dirname(input_path)
            output_path = os.path.join(output_dir, f"{base_name}_simplified.xlsx")
        
        self.generate_simplified_excel(simplified_results, output_path)
        
        return simplified_results, output_path, validation
    
    def get_summary(self, results: List[SimplifiedRow]) -> Dict[str, Any]:
        """生成化简结果的摘要 - v11"""
        snp_rows = [r for r in results if r.row_type == "SNP"]
        wt_rows = [r for r in results if r.row_type == "WT"]
        simplified_rows = [r for r in results if r.is_simplified]
        seq_ids = set(r.sequence_id for r in results)
        
        # 高亮颜色统计
        color_count = {}
        for r in snp_rows:
            if r.highlight_color:
                color_count[r.highlight_color] = color_count.get(r.highlight_color, 0) + 1
        
        # 参考序列统计
        ref_count = {}
        for r in results:
            ref_count[r.reference_name] = ref_count.get(r.reference_name, 0) + 1
        
        return {
            "total_rows": len(results),
            "total_sequence_groups": len(seq_ids),
            "groups_with_highlights": len(seq_ids) - len(simplified_rows),
            "groups_without_highlights": len(simplified_rows),
            "wt_rows": len(wt_rows),
            "snp_rows": len(snp_rows),
            "snp_with_red_bases": sum(1 for r in snp_rows if r.red_20_bases and r.red_20_bases != "-"),
            "highlight_color_distribution": color_count,
            "reference_distribution": ref_count
        }


# 便捷函数
def sort_gene_editing_file(input_path: str, output_path: str = None) -> Dict[str, Any]:
    """
    重新编号基因编辑数据文件的序号
    
    功能：
    1. 读取Excel文件
    2. 识别所有序列组（格式：0xx-refXXXX）
    3. 保持行顺序不变，只将序号依次重新编号为001, 002, 003...
    4. 生成处理后的Excel文件，保留原始格式和高亮
    
    Returns:
        Dict: 包含处理结果信息
    """
    import re
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from copy import copy
    
    # 检查文件
    if not os.path.exists(input_path):
        raise FileNotFoundError(f"文件不存在: {input_path}")
    
    # 读取源文件
    try:
        wb_src = load_workbook(input_path, data_only=False)
    except Exception as e:
        raise ValueError(f"无法读取Excel文件: {e}")
    
    ws_src = wb_src.active
    
    # 解析序列组，记录每个序号行的位置和原始序号
    sequence_pattern = re.compile(r'^(\d+)-ref(\w+)$')
    sequence_id_rows = []  # [(row_idx, orig_num, ref_name), ...] 记录序号行
    
    for row_idx, row in enumerate(ws_src.iter_rows(min_row=1), 1):
        first_cell = str(row[0].value).strip() if row[0].value else ""
        match = sequence_pattern.match(first_cell)
        
        if match:
            orig_num = match.group(1)  # 原始序号（如007）
            ref_name = match.group(2)  # 参考序列名称（如GmACC3HiTom）
            sequence_id_rows.append((row_idx, orig_num, ref_name))
    
    # 创建序号组到新序号的映射
    # 按文件中出现的顺序分配新序号，保持行顺序不变
    # 规则：连续相同的序号为一组，每组分配一个新序号
    # 例如: 007, 007, 008, 009, 007 → 001, 001, 002, 003, 004（注意最后的007是新组）
    
    # 创建新工作簿
    wb_dst = Workbook()
    ws_dst = wb_dst.active
    ws_dst.title = "重编号数据"
    
    # 创建序号行到新序号的映射（按连续组分配）
    row_to_new_seq = {}
    new_seq_counter = 0
    prev_orig_num = None
    
    for row_idx, orig_num, ref_name in sequence_id_rows:
        # 如果序号和前一行不同，则是新的序号组，分配新序号
        if orig_num != prev_orig_num:
            new_seq_counter += 1
            prev_orig_num = orig_num
        
        row_to_new_seq[row_idx] = f"{new_seq_counter:03d}-ref{ref_name}"
    
    # 复制所有数据，保持原顺序，只修改序号
    for src_row_idx, src_row in enumerate(ws_src.iter_rows(min_row=1), 1):
        for col_idx, src_cell in enumerate(src_row, 1):
            dst_cell = ws_dst.cell(row=src_row_idx, column=col_idx)
            
            # 如果是序号行的第一列，使用新序号
            if col_idx == 1 and src_row_idx in row_to_new_seq:
                dst_cell.value = row_to_new_seq[src_row_idx]
            else:
                dst_cell.value = src_cell.value
            
            # 复制样式
            if src_cell.font:
                dst_cell.font = copy(src_cell.font)
            if src_cell.fill:
                dst_cell.fill = copy(src_cell.fill)
            if src_cell.border:
                dst_cell.border = copy(src_cell.border)
            if src_cell.alignment:
                dst_cell.alignment = copy(src_cell.alignment)
            if src_cell.number_format:
                dst_cell.number_format = src_cell.number_format
    
    # 复制列宽
    for col_letter, col_dim in ws_src.column_dimensions.items():
        ws_dst.column_dimensions[col_letter].width = col_dim.width
    
    # 生成输出路径
    if not output_path:
        base, ext = os.path.splitext(input_path)
        output_path = f"{base}_renumbered{ext}"
    
    wb_dst.save(output_path)
    wb_src.close()
    wb_dst.close()
    
    return {
        "success": True,
        "output_file": output_path,
        "total_groups": new_seq_counter,
        "total_rows": len(sequence_id_rows),
        "message": f"已重新编号，共{new_seq_counter}个序号组（001-{new_seq_counter:03d}），{len(sequence_id_rows)}个序号行"
    }


def simplify_gene_editing_file(input_path: str, output_path: str = None,
                               target_bases_map: Dict[str, str] = None) -> Dict[str, Any]:
    """
    化简基因编辑数据文件 - v4.0
    
    规则：
    1. 有高亮的组：输出WT行 + 所有高亮行（完整列）
    2. 无高亮的组：仅输出WT行（只到H列）
    3. 所有行重新编号1,2,3...
    4. 20bp红色序列原样保留
    
    Returns:
        Dict: 包含结果、摘要和验证信息
    """
    processor = GeneEditingProcessor()
    results, output_file, validation = processor.process_file(input_path, output_path, target_bases_map)
    summary = processor.get_summary(results)
    
    return {
        "success": True,
        "output_file": output_file,
        "total_entries": len(results),
        "summary": summary,
        "validation": validation,
        "results": [
            {
                "original_row_num": r.original_row_num,
                "sequence_id": r.sequence_id,
                "reference_name": r.reference_name,
                "row_type": r.row_type,
                "highlight_color": r.highlight_color,
                "red_20_bases": r.red_20_bases,
                "is_highlighted": r.is_highlighted,
                "is_simplified": r.is_simplified,
                "original_cols": len(r.original_row_values) if r.original_row_values else 0
            }
            for r in results
        ]
    }
