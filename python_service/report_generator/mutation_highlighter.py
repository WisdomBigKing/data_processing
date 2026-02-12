"""
突变高亮处理器
根据基因编辑数据中的突变类型，对符合条件的行进行颜色高亮标记

筛选规则：
1. 严格匹配型（优先保留）：
   - F列和G列的突变类型完全一致（例如F列'C-T'且G列'C-T'）
   - 且H列对应位置存在该突变涉及的20bp序列

2. 兼容型（次级保留）：
   - F列或G列包含多种突变类型时（例如F列'C-T,A-T'且G列'C-T'）
   - 只要H列对应位置存在F/G列突变类型交集中任意一种的20bp序列

具体要求：
- 突变类型一致性判断：仅对F/G列中成对出现的相同突变类型进行校验（取交集）
- 多突变处理：若F/G列存在多个突变类型（用逗号分隔），需检查H列是否包含交集中任意一个突变对应的20bp序列
- H列验证：必须确保20bp序列中实际包含提示的碱基突变（如'C-T'需在20bp内找到C→T的变化）

高亮颜色：
- 黄色：C→T 突变
- 绿色：A→G 突变
- 橙色：G→A / T→C 突变

其他规则：
- 每个序号内同一突变类型只高亮一行（不重复）
- 20bp目标序列需要标红（红色字体）
"""

import os
import re
from typing import Dict, List, Any, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
from dataclasses import dataclass


# 定义高亮颜色
HIGHLIGHT_COLORS = {
    'yellow': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),  # C->T
    'green': PatternFill(start_color='92D050', end_color='92D050', fill_type='solid'),   # A->G
    'orange': PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid'),  # G->A, T->C
}

# 突变类型到颜色的映射
MUTATION_COLOR_MAP = {
    'C->T': 'yellow',
    'A->G': 'green',
    'G->A': 'orange',
    'T->C': 'orange',
}


@dataclass
class HighlightResult:
    """高亮结果"""
    row_index: int
    sequence_id: str
    mutation_type: str
    highlight_color: str
    f_value: str
    g_value: str
    # 验证结果
    verification_status: str = ""  # "correct", "error", "unknown"
    actual_mutation: str = ""  # 实际检测到的突变类型（如 "G->T"）
    verification_detail: str = ""  # 验证详情说明


class MutationHighlighter:
    """突变高亮处理器"""
    
    def __init__(self, target_sequence: str = "TGGAGACACTAGAGGGATGG"):
        """
        初始化处理器
        
        Args:
            target_sequence: 目标20bp序列，用于识别WT行
        """
        self.target_sequence = target_sequence.upper()
        self.sequence_id_pattern = re.compile(r'^(\d+)-')
        self.current_wt_sequence = ""  # 存储当前序号的WT行H列序列
    
    def _get_sequence_id(self, cell_value) -> Optional[str]:
        """从单元格值提取序号"""
        if cell_value is None:
            return None
        val = str(cell_value).strip()
        match = self.sequence_id_pattern.match(val)
        if match:
            return match.group(1)
        return None
    
    def _is_sequence_header(self, row_data: List) -> Tuple[bool, Optional[str]]:
        """检查是否为序列标题行（如 045-refGmACC1HiTom）"""
        first_cell = row_data[0] if row_data else None
        if first_cell is None:
            return False, None
        
        val = str(first_cell).strip()
        # 检查是否匹配 数字-ref 格式
        if re.match(r'^\d+-ref', val):
            seq_id = self._get_sequence_id(val)
            return True, seq_id
        return False, None
    
    def _contains_target_sequence(self, sequence: str) -> bool:
        """检查序列是否包含目标20bp序列"""
        if not sequence:
            return False
        return self.target_sequence in sequence.upper()
    
    def _find_target_sequence_position(self, sequence: str) -> Tuple[int, int]:
        """
        在序列中查找目标20bp序列的位置
        
        Returns:
            (start, end) 位置，如果未找到返回 (-1, -1)
        """
        if not sequence:
            return (-1, -1)
        
        # 在大写序列中查找目标序列位置
        upper_seq = sequence.upper()
        pos = upper_seq.find(self.target_sequence)
        if pos == -1:
            return (-1, -1)
        
        return (pos, pos + len(self.target_sequence))
    
    def _check_mutation_in_position(self, sequence: str, mutation_type: str, 
                                     target_pos: Tuple[int, int]) -> bool:
        """
        检查突变是否发生在指定的20bp目标序列位置范围内
        
        使用WT行记录的位置来检查突变行
        
        Args:
            sequence: 突变行的H列序列
            mutation_type: 突变类型 (如 "C->T")
            target_pos: WT行中目标序列的位置 (start, end)
        """
        if not sequence or not mutation_type:
            return False
        
        start, end = target_pos
        if start < 0 or end > len(sequence):
            return False
        
        # 解析突变类型 (如 "C->T")
        match = re.match(r'([ATCG])->([ATCG])', mutation_type.upper())
        if not match:
            return False
        
        mutated_base = match.group(2).lower()   # 突变后碱基（小写）
        
        # 只检查目标位置范围内是否有小写的突变碱基
        target_region = sequence[start:end]
        return mutated_base in target_region
    
    def _apply_red_font_to_20bp(self, ws, row_idx: int, col_idx: int, 
                                  sequence: str, target_pos: Tuple[int, int]):
        """
        将H列中整个20bp区域标记为红色字体
        
        Args:
            ws: 工作表
            row_idx: 行索引
            col_idx: 列索引 (H列 = 8)
            sequence: 序列字符串
            target_pos: 目标序列位置 (start, end)
        """
        if not sequence:
            return
        
        start, end = target_pos
        if start < 0 or end > len(sequence):
            return
        
        try:
            red_font = InlineFont(color='FF0000')  # 红色 - 用于20bp区域
            black_font = InlineFont(color='000000')  # 黑色 - 用于其他区域
            
            parts = []
            
            # 前缀部分（黑色）- 20bp区域之前
            if start > 0:
                parts.append(TextBlock(black_font, sequence[:start]))
            
            # 20bp区域（整个区域红色）
            parts.append(TextBlock(red_font, sequence[start:end]))
            
            # 后缀部分（黑色）- 20bp区域之后
            if end < len(sequence):
                parts.append(TextBlock(black_font, sequence[end:]))
            
            # 设置单元格为富文本
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = CellRichText(parts)
        except Exception as e:
            print(f"警告: 无法设置红色字体 (行{row_idx}): {e}")
    
    def _get_highlight_color(self, mutation_type: str) -> Optional[str]:
        """根据突变类型获取高亮颜色"""
        if not mutation_type:
            return None
        
        # 标准化突变类型格式
        normalized = mutation_type.upper().replace(' ', '')
        
        for mut_type, color in MUTATION_COLOR_MAP.items():
            if mut_type.upper() == normalized:
                return color
        
        return None
    
    def _parse_mutations_from_column(self, value: str) -> List[str]:
        """
        从F/G列值中解析所有突变类型
        
        支持格式：
        - 单个突变: "C->T" 或 "C-T"
        - 多个突变: "A-G, C-T" 或 "A->G,C->T"
        
        Returns:
            解析出的突变类型列表（标准化为 X->Y 格式）
        """
        if not value or value.strip() == '-':
            return []
        
        mutations = []
        # 按逗号或空格分割
        parts = re.split(r'[,;\s]+', value)
        
        for part in parts:
            part = part.strip().upper()
            if not part or part == '-':
                continue
            
            # 匹配突变格式: X->Y 或 X-Y
            match = re.match(r'([ATCG])[-–>]+([ATCG])', part)
            if match:
                normalized = f"{match.group(1)}->{match.group(2)}"
                mutations.append(normalized)
        
        return mutations
    
    def _find_matching_mutation(self, mutations: List[str], sequence: str, 
                                 target_pos: Tuple[int, int]) -> Optional[Tuple[str, str]]:
        """
        在突变列表中找到第一个在20bp区域内有效的突变
        
        Args:
            mutations: 突变类型列表
            sequence: H列序列
            target_pos: 目标序列位置
            
        Returns:
            (突变类型, 高亮颜色) 或 None
        """
        for mutation in mutations:
            color = self._get_highlight_color(mutation)
            if color and self._check_mutation_in_position(sequence, mutation, target_pos):
                return (mutation, color)
        return None
    
    def _get_mutation_intersection(self, f_mutations: List[str], g_mutations: List[str]) -> List[str]:
        """
        获取F列和G列突变类型的交集
        
        Args:
            f_mutations: F列解析出的突变类型列表
            g_mutations: G列解析出的突变类型列表
            
        Returns:
            交集突变类型列表
        """
        f_set = set(f_mutations)
        g_set = set(g_mutations)
        return list(f_set & g_set)
    
    def _verify_mutation_with_wt(self, mutated_sequence: str, claimed_mutation: str,
                                   target_pos: Tuple[int, int]) -> Tuple[str, str, str]:
        """
        验证突变行与WT行的实际碱基变化是否与声称的突变类型一致
        
        Args:
            mutated_sequence: 突变行的H列序列
            claimed_mutation: 声称的突变类型 (如 "C->T")
            target_pos: 目标20bp序列位置 (start, end)
            
        Returns:
            (verification_status, actual_mutation, detail)
            - verification_status: "correct" | "error" | "unknown"
            - actual_mutation: 实际检测到的突变 (如 "G->T")
            - detail: 详细说明
        """
        if not self.current_wt_sequence or not mutated_sequence:
            return ("unknown", "", "无法获取WT序列进行验证")
        
        start, end = target_pos
        if start < 0 or end > len(self.current_wt_sequence) or end > len(mutated_sequence):
            return ("unknown", "", "序列位置超出范围")
        
        # 解析声称的突变类型
        match = re.match(r'([ATCG])->([ATCG])', claimed_mutation.upper())
        if not match:
            return ("unknown", "", f"无法解析突变类型: {claimed_mutation}")
        
        claimed_from = match.group(1)  # 声称的原始碱基 (如 C)
        claimed_to = match.group(2)    # 声称的突变碱基 (如 T)
        
        # 在20bp区域内查找小写碱基（突变位置）
        wt_region = self.current_wt_sequence[start:end]
        mut_region = mutated_sequence[start:end]
        
        # 找到突变位置（小写字母位置）
        actual_mutations_found = []
        for i, (wt_char, mut_char) in enumerate(zip(wt_region, mut_region)):
            # 检查突变行是否有小写字母（表示突变）
            if mut_char.islower():
                wt_base = wt_char.upper()
                mut_base = mut_char.upper()
                actual_mutation = f"{wt_base}->{mut_base}"
                actual_mutations_found.append({
                    "position": i,
                    "wt_base": wt_base,
                    "mut_base": mut_base,
                    "actual": actual_mutation
                })
        
        if not actual_mutations_found:
            return ("unknown", "", "未在20bp区域内找到突变位置")
        
        # 检查是否有与声称突变匹配的实际突变
        for mut_info in actual_mutations_found:
            if mut_info["wt_base"] == claimed_from and mut_info["mut_base"] == claimed_to:
                return ("correct", f"{claimed_from}->{claimed_to}", 
                        f"验证通过: WT={claimed_from}, 突变后={claimed_to}")
        
        # 如果没有匹配，报告错误
        # 找出声称突变碱基(to)对应的实际WT碱基
        error_details = []
        for mut_info in actual_mutations_found:
            if mut_info["mut_base"] == claimed_to:
                # 找到了突变后碱基，但WT碱基不匹配
                actual = mut_info["actual"]
                error_details.append(
                    f"错误突变: 声称{claimed_mutation}, 实际{actual} "
                    f"(WT位置{mut_info['position']}处是{mut_info['wt_base']}而非{claimed_from})"
                )
        
        if error_details:
            # 返回第一个错误
            first_error = actual_mutations_found[0]
            return ("error", first_error["actual"], error_details[0])
        
        # 其他情况
        all_actual = ", ".join([m["actual"] for m in actual_mutations_found])
        return ("error", all_actual, f"声称{claimed_mutation}，但实际检测到: {all_actual}")
    
    def process_file(self, input_path: str, output_path: str = None) -> Dict[str, Any]:
        """
        处理Excel文件，对符合条件的行进行高亮
        
        Args:
            input_path: 输入文件路径
            output_path: 输出文件路径（可选）
            
        Returns:
            处理结果字典
        """
        # 检查文件是否存在
        if not os.path.exists(input_path):
            raise FileNotFoundError(f"文件不存在: {input_path}")
        
        # 检查文件大小
        if os.path.getsize(input_path) == 0:
            raise ValueError("文件为空")
        
        # 加载工作簿
        try:
            wb = load_workbook(input_path)
        except Exception as e:
            raise ValueError(f"无法读取Excel文件: {e}")
        
        ws = wb.active
        
        # 处理结果
        highlight_results: List[HighlightResult] = []
        current_seq_id = None
        current_has_target = False  # 当前序号的WT行是否包含目标序列
        current_target_pos = (-1, -1)  # WT行中目标序列的位置
        current_highlighted_mutations = set()  # 当前序号已高亮的突变类型（避免重复）
        
        # 遍历所有行
        for row_idx, row in enumerate(ws.iter_rows(min_row=1), 1):
            row_data = [cell.value for cell in row]
            
            # 检查是否为序列标题行
            is_header, seq_id = self._is_sequence_header(row_data)
            if is_header:
                current_seq_id = seq_id
                current_has_target = False
                current_target_pos = (-1, -1)
                current_highlighted_mutations = set()  # 新序号，重置已高亮突变类型
                continue
            
            # 跳过空行
            if all(v is None or str(v).strip() == '' for v in row_data):
                continue
            
            # 获取列值（0-indexed）
            # A=0, B=1, C=2, D=3, E=4, F=5, G=6, H=7, I=8
            d_value = str(row_data[3]).strip() if len(row_data) > 3 and row_data[3] else ""
            e_value = str(row_data[4]).strip() if len(row_data) > 4 and row_data[4] else ""
            f_value = str(row_data[5]).strip() if len(row_data) > 5 and row_data[5] else ""
            g_value = str(row_data[6]).strip() if len(row_data) > 6 and row_data[6] else ""
            h_value = str(row_data[7]).strip() if len(row_data) > 7 and row_data[7] else ""
            
            # 如果是WT行，检查是否包含目标序列，并记录位置和WT序列
            if d_value == 'WT' and e_value == 'WT':
                pos = self._find_target_sequence_position(h_value)
                if pos[0] >= 0:
                    current_has_target = True
                    current_target_pos = pos
                    self.current_wt_sequence = h_value  # 保存WT的H列序列用于验证
                continue
            
            # 如果当前序号的WT行不包含目标序列，跳过
            if not current_has_target:
                continue
            
            # 检查F列和G列是否有值
            if f_value in ['-', ''] and g_value in ['-', '']:
                continue
            
            # 解析F列和G列中的所有突变类型
            f_mutations = self._parse_mutations_from_column(f_value)
            g_mutations = self._parse_mutations_from_column(g_value)
            
            # 如果F或G列都没有有效突变，跳过
            if not f_mutations and not g_mutations:
                continue
            
            # 获取F/G列突变类型的交集
            # 规则：只有在F和G列都出现的突变类型才进行验证
            common_mutations = self._get_mutation_intersection(f_mutations, g_mutations)
            
            # 如果没有交集（F和G没有共同的突变类型），跳过
            if not common_mutations:
                continue
            
            # 在交集中找到第一个在20bp区域内有效且未被高亮的突变
            matched_mutation = None
            highlight_color = None
            
            for mutation in common_mutations:
                # 检查该突变类型是否已在当前序号内高亮过
                if mutation in current_highlighted_mutations:
                    continue
                
                # 检查该突变是否在20bp区域内（H列验证）
                color = self._get_highlight_color(mutation)
                if color and self._check_mutation_in_position(h_value, mutation, current_target_pos):
                    matched_mutation = mutation
                    highlight_color = color
                    break
            
            if not matched_mutation or not highlight_color:
                continue
            
            # 验证突变：与WT的20bp对比，确保实际突变与声称的一致
            verify_status, actual_mut, verify_detail = self._verify_mutation_with_wt(
                h_value, matched_mutation, current_target_pos
            )
            
            # 【关键】只有验证通过才应用高亮，错误突变（如G-T被声称为C-T）不高亮
            if verify_status == "error":
                print(f"[跳过] 行{row_idx}: 验证失败 - {verify_detail}")
                continue
            
            # 应用高亮（仅验证通过的行）
            fill = HIGHLIGHT_COLORS.get(highlight_color)
            if fill:
                for cell in row:
                    cell.fill = fill
                
                # 将H列中整个20bp区域标红
                self._apply_red_font_to_20bp(ws, row_idx, 8, h_value, current_target_pos)
                
                # 记录已高亮的突变类型（避免同一序号内重复）
                current_highlighted_mutations.add(matched_mutation)
                
                highlight_results.append(HighlightResult(
                    row_index=row_idx,
                    sequence_id=current_seq_id or "",
                    mutation_type=matched_mutation,
                    highlight_color=highlight_color,
                    f_value=f_value,
                    g_value=g_value,
                    verification_status=verify_status,
                    actual_mutation=actual_mut,
                    verification_detail=verify_detail
                ))
        
        # 生成输出路径
        if not output_path:
            base, ext = os.path.splitext(input_path)
            output_path = f"{base}_highlighted{ext}"
        
        # 保存文件
        wb.save(output_path)
        
        # 统计结果
        color_stats = {}
        mutation_stats = {}
        verification_stats = {"correct": 0, "error": 0, "unknown": 0}
        error_rows = []  # 记录验证失败的行
        
        for result in highlight_results:
            color_stats[result.highlight_color] = color_stats.get(result.highlight_color, 0) + 1
            mutation_stats[result.mutation_type] = mutation_stats.get(result.mutation_type, 0) + 1
            
            # 统计验证结果
            if result.verification_status in verification_stats:
                verification_stats[result.verification_status] += 1
            
            # 记录错误突变
            if result.verification_status == "error":
                error_rows.append({
                    "row": result.row_index,
                    "sequence_id": result.sequence_id,
                    "claimed_mutation": result.mutation_type,
                    "actual_mutation": result.actual_mutation,
                    "detail": result.verification_detail
                })
        
        return {
            "success": True,
            "output_file": output_path,
            "total_highlighted": len(highlight_results),
            "target_sequence": self.target_sequence,
            "color_statistics": color_stats,
            "mutation_statistics": mutation_stats,
            "verification_statistics": verification_stats,
            "error_mutations": error_rows,
            "highlighted_rows": [
                {
                    "row": r.row_index,
                    "sequence_id": r.sequence_id,
                    "mutation_type": r.mutation_type,
                    "color": r.highlight_color,
                    "verification": r.verification_status,
                    "actual_mutation": r.actual_mutation,
                    "verification_detail": r.verification_detail
                }
                for r in highlight_results
            ]
        }


def highlight_mutations(input_path: str, output_path: str = None,
                       target_sequence: str = "TGGAGACACTAGAGGGATGG") -> Dict[str, Any]:
    """
    对基因编辑数据文件进行突变高亮处理
    
    Args:
        input_path: 输入文件路径
        output_path: 输出文件路径（可选）
        target_sequence: 目标20bp序列
        
    Returns:
        处理结果字典
    """
    highlighter = MutationHighlighter(target_sequence)
    return highlighter.process_file(input_path, output_path)


if __name__ == "__main__":
    # 测试
    import sys
    if len(sys.argv) > 1:
        result = highlight_mutations(sys.argv[1])
        print(f"处理完成: {result}")
